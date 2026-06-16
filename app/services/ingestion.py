import hashlib
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

import httpx
from openpyxl import load_workbook

from app.core.config import Settings


@dataclass(frozen=True)
class DownloadResult:
    path: Path
    checksum: str
    byte_size: int
    etag: str | None
    last_modified: str | None


@dataclass(frozen=True)
class ParsedCell:
    department_code: str
    department_name: str
    municipality_code: str
    municipality_name: str
    year: int
    area: str
    sex: str
    age: int
    population: int


class DANEPopulationConnector:
    SHEET = "PobMunicipalxÁreaSexoEdad"
    HEADER_ROW = 9
    AGE_PATTERN = re.compile(r"^(Hombres|Mujeres) (\d{1,3}) años?(?: y más)?$")

    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    def download(self, url: str, source_id: str) -> DownloadResult:
        self.settings.raw_storage_path.mkdir(parents=True, exist_ok=True)
        temporary = self.settings.raw_storage_path / f".{source_id}.download"
        digest = hashlib.sha256()
        total = 0
        with httpx.stream(
            "GET",
            url,
            timeout=httpx.Timeout(self.settings.request_timeout_seconds, read=120.0),
            follow_redirects=True,
            headers={"User-Agent": "Databolico-GenZ-API-V3/3.0"},
        ) as response:
            response.raise_for_status()
            with temporary.open("wb") as output:
                for chunk in response.iter_bytes():
                    total += len(chunk)
                    if total > self.settings.source_max_bytes:
                        raise ValueError("Source exceeds configured maximum size")
                    digest.update(chunk)
                    output.write(chunk)
            checksum = digest.hexdigest()
            destination_dir = self.settings.raw_storage_path / source_id / checksum
            destination_dir.mkdir(parents=True, exist_ok=True)
            destination = destination_dir / "source.xlsx"
            if destination.exists():
                temporary.unlink(missing_ok=True)
            else:
                shutil.move(temporary, destination)
            return DownloadResult(
                destination,
                checksum,
                total,
                response.headers.get("etag"),
                response.headers.get("last-modified"),
            )

    def _age_columns(self, sheet) -> list[tuple[int, str, int]]:
        """Locate the (column index, sex, age) triples from the verified header row."""
        header = next(sheet.iter_rows(min_row=self.HEADER_ROW, max_row=self.HEADER_ROW, values_only=True))
        age_columns: list[tuple[int, str, int]] = []
        for index, value in enumerate(header):
            if not isinstance(value, str):
                continue
            match = self.AGE_PATTERN.match(value.strip())
            if match:
                age_columns.append((index, "M" if match.group(1) == "Hombres" else "F", int(match.group(2))))
        # One column per (sex, age) for ages 0..max_age inclusive, both sexes.
        expected = (self.settings.max_age + 1) * 2
        if len(age_columns) != expected:
            raise ValueError(
                f"Expected {expected} sex-age columns for max_age={self.settings.max_age}, "
                f"found {len(age_columns)}"
            )
        return age_columns

    def iter_cells(
        self,
        path: Path,
        municipality_codes: set[str],
        years: set[int] | None = None,
    ):
        """Stream the source one source-row at a time.

        Yields ``(row_cells, control_key, control_value)`` per accepted (municipality,
        year) row, where ``row_cells`` is the list of sex×age :class:`ParsedCell` for
        that row only. This keeps peak memory bounded to a single municipality-year
        (~202 cells) instead of materializing the whole country (~5.6M cells at once),
        which is what was OOM-killing the container on the shared VPS.
        """
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if self.SHEET not in workbook.sheetnames:
                raise ValueError(f"Required sheet {self.SHEET!r} is missing")
            sheet = workbook[self.SHEET]
            age_columns = self._age_columns(sheet)
            # An empty target set means "every municipality in the file" (all of Colombia).
            select_all = not municipality_codes
            for row in sheet.iter_rows(min_row=self.HEADER_ROW + 1, values_only=True):
                if row[2] is None:
                    continue
                municipality_code = str(row[2]).zfill(5)
                year = int(row[4])
                area = str(row[5]).strip()
                if area != "Total":
                    continue
                if select_all:
                    # Defensive: skip department/national aggregate rows (codes ending in 000);
                    # real DIVIPOLA municipality codes never do.
                    if municipality_code.endswith("000"):
                        continue
                elif municipality_code not in municipality_codes:
                    continue
                if years is not None and year not in years:
                    continue
                control_key = (municipality_code, year, area)
                control_value = {"total": int(row[6]), "M": int(row[7]), "F": int(row[8])}
                department_code = str(row[0]).zfill(2)
                department_name = str(row[1]).strip()
                municipality_name = str(row[3]).strip()
                row_cells = [
                    ParsedCell(
                        department_code=department_code,
                        department_name=department_name,
                        municipality_code=municipality_code,
                        municipality_name=municipality_name,
                        year=year,
                        area=area,
                        sex=sex,
                        age=age,
                        population=int(row[column]),
                    )
                    for column, sex, age in age_columns
                ]
                yield row_cells, control_key, control_value
        finally:
            workbook.close()

    def parse(
        self,
        path: Path,
        municipality_codes: set[str],
        years: set[int] | None = None,
    ) -> tuple[list[ParsedCell], dict]:
        """Eagerly materialize every cell. Convenience wrapper over :meth:`iter_cells`
        for tests and small inputs; production ingestion streams via ``iter_cells``."""
        cells: list[ParsedCell] = []
        control_totals: dict[tuple[str, int, str], dict] = {}
        for row_cells, control_key, control_value in self.iter_cells(path, municipality_codes, years):
            cells.extend(row_cells)
            control_totals[control_key] = control_value
        if not cells:
            raise ValueError("No rows matched the configured municipalities and years")
        return cells, control_totals
